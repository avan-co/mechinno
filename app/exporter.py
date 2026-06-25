from __future__ import annotations

from io import BytesIO
import sqlite3
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GOLD = "C9A44C"
INK = "141414"
DARK = "111827"
LIGHT = "F8F5EE"
BORDER = Side(style="thin", color="D8C894")


REPORTS = {
    "members": {
        "title": "اعضا",
        "query": """
            SELECT code, full_name, team_name, desks, lockers, power_strips,
                   phone, national_id, notes
            FROM members
            ORDER BY id
        """,
        "headers": ["کد", "نام", "تیم/شرکت", "میز", "کمد", "سه‌راهی", "تماس", "کدملی", "توضیحات"],
    },
    "teams": {
        "title": "تیم‌ها",
        "query": """
            SELECT name, leader, phone, desk_count, lockers, power_strips,
                   joined_at, warning, notes
            FROM teams
            ORDER BY id
        """,
        "headers": ["نام تیم", "سرگروه", "تماس", "تعداد میز", "کمد", "سه‌راهی", "عضویت", "اخطار", "توضیحات"],
    },
    "lockers": {
        "title": "کمدها",
        "query": """
            SELECT locker_number, status, assigned_to, delivered_at, key_number, spare_key, notes
            FROM lockers
            ORDER BY locker_number
        """,
        "headers": ["شماره کمد", "وضعیت", "اختصاص به", "تاریخ تحویل", "شماره کلید", "کلید یدک", "توضیحات"],
    },
    "charges": {
        "title": "شارژ و اجاره",
        "query": """
            SELECT fiscal_year, team_name, leader, desk_count, month_name,
                   amount, note, charge_rate, rent_rate
            FROM charges
            ORDER BY fiscal_year, team_name, month_index
        """,
        "headers": ["سال", "تیم", "سرگروه", "میز", "ماه", "مبلغ", "یادداشت", "نرخ شارژ", "نرخ اجاره"],
    },
    "transactions": {
        "title": "مالی",
        "query": """
            SELECT b.sheet_name, t.tx_date, t.description, t.amount, t.category,
                   t.notes, t.suspected_amount_note, b.petty_cash_holder
            FROM transactions t
            JOIN financial_batches b ON b.id = t.batch_id
            ORDER BY b.id, t.id
        """,
        "headers": ["دوره", "تاریخ", "شرح", "مبلغ", "دسته", "توضیحات", "مبلغ مشکوک در توضیحات", "دارنده تنخواه"],
    },
    "plans": {
        "title": "برنامه‌ها",
        "query": """
            SELECT plan_number, status, title, proposed_budget, cost_type, schedule, notes
            FROM plans
            ORDER BY plan_number
        """,
        "headers": ["شماره", "وضعیت", "عنوان", "بودجه پیشنهادی", "نوع هزینه", "زمان‌بندی", "توضیحات"],
    },
    "warnings": {
        "title": "هشدارهای داده",
        "query": """
            SELECT file_name, sheet_name, row_number, message, payload
            FROM import_warnings
            ORDER BY id
        """,
        "headers": ["فایل", "شیت", "ردیف", "پیام", "جزئیات"],
    },
}


def _style_sheet(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3

    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=GOLD)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)
            if cell.row % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    for column_cells in ws.columns:
        max_length = 8
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 42)


def _add_report_sheet(wb: Workbook, title: str, headers: list[str], rows: Iterable[sqlite3.Row]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws.append(headers)
    for row in rows:
        ws.append([row[index] for index in range(len(headers))])
    _style_sheet(ws)


def build_workbook(conn: sqlite3.Connection, report_key: str = "all") -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    keys = list(REPORTS.keys()) if report_key == "all" else [report_key]
    for key in keys:
        if key not in REPORTS:
            raise KeyError(key)
        report = REPORTS[key]
        rows = conn.execute(report["query"]).fetchall()
        _add_report_sheet(wb, report["title"], report["headers"], rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
