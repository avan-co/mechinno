from __future__ import annotations

import json
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .database import init_db, reset_imported_data


MONTHS = [
    ("E", 1, "فروردین"),
    ("F", 2, "اردیبهشت"),
    ("G", 3, "خرداد"),
    ("H", 4, "تیر"),
    ("I", 5, "مرداد"),
    ("J", 6, "شهریور"),
    ("K", 7, "مهر"),
    ("L", 8, "آبان"),
    ("M", 9, "آذر"),
    ("N", 10, "دی"),
    ("O", 11, "بهمن"),
    ("P", 12, "اسفند"),
]

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")


@dataclass(frozen=True)
class WorkbookSource:
    path: Path
    name: str


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).translate(PERSIAN_DIGITS).strip()


def parse_int(value: Any) -> int | None:
    text = clean(value)
    if not text or text == "-":
        return None
    text = text.replace(",", "").replace("٬", "")
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return int(float(text))
    return None


def parse_amount(value: Any) -> int | None:
    return parse_int(value)


def parse_jalali_date(value: Any) -> str:
    text = clean(value)
    if not text or text == "-":
        return ""
    match = re.fullmatch(r"(\d{2,4})/(\d{1,2})/(\d{1,2})", text)
    if not match:
        return text
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}/{month:02d}/{day:02d}"


def compose_jalali_date(day: Any, month: Any, year: Any) -> str:
    d = parse_int(day)
    m = parse_int(month)
    y = parse_int(year)
    if d is None or m is None or y is None:
        return ""
    return f"{y:04d}/{m:02d}/{d:02d}"


def is_valid_jalali_date(date_text: str) -> bool:
    match = re.fullmatch(r"(\d{4})/(\d{2})/(\d{2})", date_text)
    if not match:
        return True
    _, month, day = (int(part) for part in match.groups())
    if month < 1 or month > 12:
        return False
    max_day = 31 if month <= 6 else 30
    if month == 12:
        max_day = 29
    return 1 <= day <= max_day


def cell_reader(sheet: Worksheet):
    merged_lookup: dict[str, str] = {}
    for merged_range in sheet.merged_cells.ranges:
        anchor = merged_range.start_cell.coordinate
        for row in sheet[merged_range.coord]:
            for cell in row:
                merged_lookup[cell.coordinate] = anchor

    def value(coordinate: str) -> Any:
        anchor = merged_lookup.get(coordinate, coordinate)
        return sheet[anchor].value

    return value


def warn(
    conn: sqlite3.Connection,
    file_name: str,
    sheet_name: str,
    row_number: int | None,
    message: str,
    payload: dict[str, Any] | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO import_warnings (file_name, sheet_name, row_number, message, payload)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            file_name,
            sheet_name,
            row_number,
            message,
            json.dumps(payload or {}, ensure_ascii=False),
        ),
    )


def import_innovation_center(conn: sqlite3.Connection, source: WorkbookSource) -> None:
    workbook = load_workbook(source.path, data_only=True)

    members_sheet = workbook["Members"]
    get = cell_reader(members_sheet)
    for row in range(4, members_sheet.max_row + 1):
        full_name = clean(get(f"C{row}"))
        if not full_name:
            continue
        conn.execute(
            """
            INSERT INTO members
                (row_number, code, full_name, team_name, desks, lockers, power_strips,
                 phone, national_id, notes, source_file, source_sheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row,
                clean(get(f"B{row}")),
                full_name,
                clean(get(f"D{row}")),
                clean(get(f"E{row}")),
                clean(get(f"F{row}")),
                clean(get(f"G{row}")),
                clean(get(f"H{row}")),
                clean(get(f"I{row}")),
                clean(get(f"J{row}")),
                source.name,
                members_sheet.title,
            ),
        )

    teams_sheet = workbook["Teams"]
    get = cell_reader(teams_sheet)
    for row in range(6, 20):
        name = clean(get(f"B{row}"))
        leader = clean(get(f"C{row}"))
        if not name and not leader:
            continue
        conn.execute(
            """
            INSERT INTO teams
                (row_number, name, leader, phone, desk_count, lockers, power_strips,
                 joined_at, warning, notes, source_file, source_sheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row,
                name,
                leader,
                clean(get(f"D{row}")),
                parse_int(get(f"E{row}")),
                clean(get(f"F{row}")),
                clean(get(f"G{row}")),
                parse_jalali_date(get(f"H{row}")),
                clean(get(f"I{row}")),
                clean(get(f"J{row}")),
                source.name,
                teams_sheet.title,
            ),
        )

    lockers_sheet = workbook["lockers"]
    get = cell_reader(lockers_sheet)
    for row in range(6, lockers_sheet.max_row + 1):
        locker_number = parse_int(get(f"A{row}"))
        if locker_number is None:
            continue
        delivered_at = parse_jalali_date(get(f"D{row}"))
        if delivered_at and not is_valid_jalali_date(delivered_at):
            warn(
                conn,
                source.name,
                lockers_sheet.title,
                row,
                "تاریخ تحویل کمد معتبر به نظر نمی‌رسد.",
                {"delivered_at": delivered_at, "locker_number": locker_number},
            )
        conn.execute(
            """
            INSERT INTO lockers
                (locker_number, status, assigned_to, delivered_at, key_number,
                 spare_key, notes, source_file, source_sheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                locker_number,
                clean(get(f"B{row}")),
                clean(get(f"C{row}")),
                delivered_at,
                clean(get(f"E{row}")),
                clean(get(f"F{row}")),
                "",
                source.name,
                lockers_sheet.title,
            ),
        )

    plans_sheet = workbook["plans"]
    get = cell_reader(plans_sheet)
    for row in range(6, plans_sheet.max_row + 1):
        plan_number = parse_int(get(f"A{row}"))
        title = clean(get(f"C{row}"))
        if plan_number is None and not title:
            continue
        conn.execute(
            """
            INSERT INTO plans
                (plan_number, status, title, proposed_budget, cost_type, schedule,
                 notes, source_file, source_sheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                plan_number,
                clean(get(f"B{row}")),
                title,
                parse_amount(get(f"D{row}")),
                clean(get(f"E{row}")),
                clean(get(f"F{row}")),
                "",
                source.name,
                plans_sheet.title,
            ),
        )


def import_charges(conn: sqlite3.Connection, source: WorkbookSource) -> None:
    workbook = load_workbook(source.path, data_only=True)
    for sheet in workbook.worksheets:
        get = cell_reader(sheet)
        charge_rate = parse_amount(get("T1"))
        rent_rate = parse_amount(get("T3"))
        for row in range(6, sheet.max_row + 1):
            team_name = clean(get(f"B{row}"))
            leader = clean(get(f"C{row}"))
            if not team_name and not leader:
                continue
            if clean(get(f"A{row}")) and not parse_int(get(f"A{row}")):
                continue
            note = clean(get(f"Q{row}"))
            for column, month_index, month_name in MONTHS:
                raw_amount = get(f"{column}{row}")
                amount = parse_amount(raw_amount)
                if amount is None:
                    continue
                conn.execute(
                    """
                    INSERT INTO charges
                        (fiscal_year, team_name, leader, desk_count, month_index,
                         month_name, amount, note, charge_rate, rent_rate,
                         source_file, source_sheet)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sheet.title,
                        team_name,
                        leader,
                        parse_int(get(f"D{row}")),
                        month_index,
                        month_name,
                        amount,
                        note,
                        charge_rate,
                        rent_rate,
                        source.name,
                        sheet.title,
                    ),
                )


def categorize_transaction(description: str, amount: int | None) -> str:
    if amount is None:
        return "نامشخص"
    if amount < 0:
        return "هزینه"
    keywords = ("آبونمان", "شارژ", "سود", "سهم مرکز", "واریز")
    if any(keyword in description for keyword in keywords):
        return "درآمد"
    return "دریافت"


def import_finance(conn: sqlite3.Connection, source: WorkbookSource) -> None:
    workbook = load_workbook(source.path, data_only=True)
    for sheet in workbook.worksheets:
        get = cell_reader(sheet)
        conn.execute(
            """
            INSERT INTO financial_batches
                (sheet_name, petty_cash_holder, petty_cash_number, previous_balance,
                 new_deposit, total_balance, received_at, from_date, to_date, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sheet.title,
                clean(get("I1")),
                clean(get("K1")),
                parse_amount(get("I2")),
                parse_amount(get("I3")),
                parse_amount(get("I4")),
                parse_jalali_date(get("K2")),
                parse_jalali_date(get("K3")),
                parse_jalali_date(get("K4")),
                source.name,
            ),
        )
        batch_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        for row in range(6, 28):
            row_number = clean(get(f"A{row}"))
            description = clean(get(f"F{row}"))
            amount = parse_amount(get(f"I{row}"))
            notes = clean(get(f"J{row}"))
            suspected_amount_note = None
            if row_number and (description or amount is not None or notes):
                note_amount = parse_amount(notes)
                if note_amount is not None and (amount is None or amount == 0):
                    suspected_amount_note = note_amount
                    warn(
                        conn,
                        source.name,
                        sheet.title,
                        row,
                        "یک مبلغ احتمالی در ستون توضیحات پیدا شد.",
                        {
                            "description": description,
                            "amount_column": amount,
                            "notes_column": notes,
                        },
                    )
                conn.execute(
                    """
                    INSERT INTO transactions
                        (batch_id, row_number, invoice_count, tx_date, description,
                         amount, notes, category, suspected_amount_note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        parse_int(row_number),
                        clean(get(f"B{row}")),
                        compose_jalali_date(get(f"C{row}"), get(f"D{row}"), get(f"E{row}")),
                        description,
                        amount,
                        notes,
                        categorize_transaction(description, amount),
                        suspected_amount_note,
                    ),
                )


def import_all(conn: sqlite3.Connection, base_dir: str | Path) -> None:
    base = Path(base_dir)
    init_db(conn)
    reset_imported_data(conn)
    sources = {
        "innovation": WorkbookSource(base / "Innovation Center.xlsx", "Innovation Center.xlsx"),
        "charges": WorkbookSource(base / "CHARGE.xlsx", "CHARGE.xlsx"),
        "finance": WorkbookSource(base / "finance.xlsx", "finance.xlsx"),
    }
    for source in sources.values():
        if not source.path.exists():
            raise FileNotFoundError(f"Required workbook not found: {source.path}")

    import_innovation_center(conn, sources["innovation"])
    import_charges(conn, sources["charges"])
    import_finance(conn, sources["finance"])
    conn.execute(
        "INSERT INTO import_runs (source_files) VALUES (?)",
        (json.dumps([source.name for source in sources.values()], ensure_ascii=False),),
    )
    conn.commit()
